import sys
import os
import openpyxl
import cv2 as cv
def image_num(path):
    for k in big_list:
      sys.stdout=open(path,'a')
      print(k)

def self_get_imlist(path):
    """all return a list of filenames for all jpg images in a directory """
    return [os.path.join(path, f) for f in tsnefile_list if f.endswith('.jpg')]

path_dir='./resized_16000_0.2/'
file_list=os.listdir(path_dir)

big_list=[]

for k in file_list:
    newstr=k.replace(".jpg","")
    newstr=int(newstr)
    big_list.append(newstr)

print(big_list)
image_num('norm_16000_delete_each.txt')


wb = openpyxl.load_workbook('RGB_최종_DBSCAN.xlsx')

sheet1=wb.active

cluter_list=[]
for i in range (15813):
    for a in range(114):
      if sheet1.cell(row=i+1,column=2).value == a:
         file=open("tsen_%i.txt" %a,'a')
         a=str(sheet1.cell(row=1 + i, column=1).value)
         file.write('%s \n'%a.zfill(5))



for i in range(114):
    os.mkdir('./tsne/tsne_%i/'%i)

# 파일 불러오기
num=0
for i in range(114):
 f = open("tsen_%i.txt"%i, mode='rt')
 lines = f.readlines()
 print(lines)
 tsnefile_list=[0]*len(lines)
 for line, i in zip(lines,range(0,len(lines))):
     line=line.replace(" \n","")
     tsnefile_list[i]= str(line) +".jpg"
 image=self_get_imlist('./resized_16000_0.2/')
 print(tsnefile_list)
 print(image)
 for j in image:
    img=cv.imread(j)
    a=j
    a=a.replace('./resized_16000_0.2/','')
    cv.imwrite(os.path.join('./tsne/tsne_%i/'%num,a),img)
 num=num+1